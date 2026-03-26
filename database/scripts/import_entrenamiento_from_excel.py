#!/usr/bin/env python
from __future__ import annotations

import argparse
import datetime as dt
import fnmatch
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


ROOT = Path(__file__).resolve().parents[2]
ENV_PATH = ROOT / ".env"
FORM_PATH = ROOT / "app" / "Views" / "entrenamiento" / "form.php"
DEFAULT_PATTERN = "Plan de Entrenamiento*.xlsx"


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    text = text.replace("–", "-").replace("—", "-").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def timestamp_to_string(value: Any, fallback: Any = None) -> str:
    if isinstance(value, dt.datetime):
        return value.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min).strftime("%Y-%m-%d %H:%M:%S")
    if fallback is not None:
        return timestamp_to_string(fallback)

    text = clean_text(value)
    if text == "":
        raise ValueError("Marca temporal vacía.")

    return dt.datetime.fromisoformat(text).replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def parse_php_string_list(var_name: str, text: str) -> list[str]:
    match = re.search(rf"\${var_name}\s*=\s*\[(.*?)\];", text, re.S)
    if not match:
        return []

    block = match.group(1)
    values: list[str] = []
    for raw in re.findall(r"'((?:\\'|[^'])*)'", block):
        values.append(raw.replace("\\'", "'"))

    return values


def load_topic_options(form_path: Path) -> dict[str, list[str]]:
    text = form_path.read_text(encoding="utf-8", errors="replace")
    return {
        "suicidio": parse_php_string_list("suicidioOptions", text),
        "violencias": parse_php_string_list("violenciasOptions", text),
        "adicciones": parse_php_string_list("adiccionesOptions", text),
        "otros_temas_salud_mental": parse_php_string_list("otrosTemasOptions", text),
    }


def find_column(headers: list[Any], prefix: str) -> int:
    for idx, header in enumerate(headers):
        if isinstance(header, str) and normalize_text(header).startswith(normalize_text(prefix)):
            return idx
    return -1


def find_columns(headers: list[Any], prefix: str) -> list[int]:
    indices: list[int] = []
    for idx, header in enumerate(headers):
        if isinstance(header, str) and normalize_text(header).startswith(normalize_text(prefix)):
            indices.append(idx)
    return indices


def first_non_empty(values: list[Any]) -> str:
    for value in values:
        text = clean_text(value)
        if text != "":
            return text
    return ""


def get_cell(row: tuple[Any, ...], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def load_name_aliases(path: str | None) -> dict[str, str]:
    if not path:
        return {}

    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("El archivo de aliases debe ser un JSON objeto {excel_name: target}.")

    aliases: dict[str, str] = {}
    for key, value in raw.items():
        aliases[normalize_text(key)] = str(value)
    return aliases


def build_user_indexes(
    connection: pymysql.connections.Connection,
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    by_email: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = {}

    with connection.cursor() as cur:
        cur.execute("SELECT id, name, email FROM users")
        for user_id, name, email in cur.fetchall():
            row = {"id": int(user_id), "name": str(name), "email": str(email)}
            by_email[normalize_text(email)] = row
            by_name.setdefault(normalize_text(name), []).append(row)

    return by_email, by_name


def resolve_user(
    excel_name: str,
    aliases: dict[str, str],
    by_email: dict[str, dict[str, Any]],
    by_name: dict[str, list[dict[str, Any]]],
) -> tuple[dict[str, Any] | None, str]:
    normalized_name = normalize_text(excel_name)
    target = aliases.get(normalized_name)
    if target:
        target_normalized = normalize_text(target)
        if target_normalized in by_email:
            return by_email[target_normalized], "alias_email"
        matches = by_name.get(target_normalized, [])
        if len(matches) == 1:
            return matches[0], "alias_name"

    matches = by_name.get(normalized_name, [])
    if len(matches) == 1:
        return matches[0], "name"
    if len(matches) > 1:
        return None, "ambiguous"
    return None, "missing"


def extract_topics(cell_value: Any, allowed_topics: list[str]) -> list[str]:
    text = normalize_text(cell_value)
    if text == "":
        return []

    hits: list[str] = []
    for topic in allowed_topics:
        if normalize_text(topic) in text:
            hits.append(topic)

    return hits


def load_existing_keys(connection: pymysql.connections.Connection) -> set[tuple[str, ...]]:
    keys: set[tuple[str, ...]] = set()
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT user_id, subregion, municipality, created_at, payload
            FROM entrenamiento_plans
            """
        )
        for user_id, subregion, municipality, created_at, payload_json in cur.fetchall():
            payload = json.loads(payload_json) if payload_json else {}
            keys.add(
                (
                    str(int(user_id)),
                    normalize_text(subregion),
                    normalize_text(municipality),
                    clean_text(created_at),
                    json.dumps(payload, ensure_ascii=False, sort_keys=True),
                )
            )
    return keys


def collect_files(directory: str, pattern: str) -> list[str]:
    normalized_pattern = normalize_text(pattern)
    files = [
        os.path.join(directory, name)
        for name in os.listdir(directory)
        if fnmatch.fnmatch(normalize_text(name), normalized_pattern)
    ]
    return sorted(files)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa masivamente planes de entrenamiento desde Excel a entrenamiento_plans."
    )
    parser.add_argument("--file", help="Ruta exacta del xlsx a importar.")
    parser.add_argument("--dir", default=str(Path.home() / "Downloads"), help="Carpeta donde están los xlsx.")
    parser.add_argument("--pattern", default=DEFAULT_PATTERN, help="Patrón de archivos a importar.")
    parser.add_argument("--editable", type=int, default=1, choices=[0, 1], help="Valor del campo editable.")
    parser.add_argument("--name-aliases", help="JSON opcional con aliases de nombre o email.")
    parser.add_argument(
        "--skip-unresolved",
        action="store_true",
        help="Importa las filas válidas aunque queden profesionales sin resolver.",
    )
    parser.add_argument("--commit", action="store_true", help="Inserta en BD. Sin esto solo hace dry-run.")
    return parser.parse_args()


def resolve_input_files(args: argparse.Namespace) -> list[str]:
    if args.file:
        return [args.file]
    return collect_files(args.dir, args.pattern)


def main() -> int:
    args = parse_args()
    files = resolve_input_files(args)
    if not files:
        print("No se encontraron archivos para importar.", file=sys.stderr)
        return 1

    env = load_env(ENV_PATH)
    topic_options = load_topic_options(FORM_PATH)
    aliases = load_name_aliases(args.name_aliases)

    connection = pymysql.connect(
        host=env["DB_HOST"],
        port=int(env.get("DB_PORT", 3306)),
        user=env["DB_USERNAME"],
        password=env.get("DB_PASSWORD", ""),
        database=env["DB_DATABASE"],
        charset="utf8mb4",
        autocommit=False,
    )

    stats = {
        "files": len(files),
        "rows_seen": 0,
        "rows_ready": 0,
        "rows_inserted": 0,
        "rows_skipped_existing": 0,
        "matched_by_name": 0,
        "matched_by_alias_name": 0,
        "matched_by_alias_email": 0,
        "warnings": 0,
        "unresolved": 0,
    }
    warnings: list[str] = []
    unresolved: list[str] = []

    try:
        by_email, by_name = build_user_indexes(connection)
        existing_keys = load_existing_keys(connection)

        insert_sql = """
            INSERT INTO entrenamiento_plans (
                user_id,
                professional_name,
                professional_email,
                subregion,
                municipality,
                editable,
                payload,
                created_at,
                updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        with connection.cursor() as cur:
            for file_path in files:
                filename = os.path.basename(file_path)
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                headers = list(next(rows))

                name_col = find_column(headers, "Seleccione su nombre")
                subregion_col = find_column(headers, "Seleccione la subregión")
                municipality_cols = find_columns(headers, "Seleccione el municipio")
                suicidio_col = find_column(headers, "SUICIDIO")
                violencias_col = find_column(headers, "VIOLENCIAS")
                adicciones_col = find_column(headers, "ADICCIONES")
                otros_col = find_column(headers, "OTROS TEMAS DE INTERÉS EN SALUD MENTAL")
                tema_1_col = find_column(headers, "1 Tema propuesto")
                tema_2_col = find_column(headers, "2 Tema propuesto")
                tema_3_col = find_column(headers, "3 Tema propuesto")
                tema_4_col = find_column(headers, "4 Tema propuesto")
                justificacion_col = find_column(headers, "Breve justificación")

                required_columns = [
                    name_col,
                    subregion_col,
                    suicidio_col,
                    violencias_col,
                    adicciones_col,
                    otros_col,
                    tema_1_col,
                    tema_2_col,
                    tema_3_col,
                    tema_4_col,
                    justificacion_col,
                ]
                if municipality_cols == [] or any(index < 0 for index in required_columns):
                    unresolved.append(f"{filename}: no se pudieron identificar todas las columnas esperadas.")
                    continue

                for row_number, row in enumerate(rows, start=2):
                    stats["rows_seen"] += 1
                    row = tuple(row)

                    excel_name = clean_text(get_cell(row, name_col))
                    user, mode = resolve_user(excel_name, aliases, by_email, by_name)
                    if user is None:
                        unresolved.append(
                            f"{filename} fila {row_number}: no se encontró usuario para {excel_name!r} ({mode})."
                        )
                        continue
                    stats[f"matched_by_{mode}"] += 1

                    subregion = first_non_empty([get_cell(row, subregion_col)]).upper()
                    municipality = first_non_empty([get_cell(row, idx) for idx in municipality_cols]).upper()
                    if subregion == "" or municipality == "":
                        unresolved.append(f"{filename} fila {row_number}: subregión o municipio vacíos.")
                        continue

                    suicidio = extract_topics(get_cell(row, suicidio_col), topic_options["suicidio"])
                    violencias = extract_topics(get_cell(row, violencias_col), topic_options["violencias"])
                    adicciones = extract_topics(get_cell(row, adicciones_col), topic_options["adicciones"])
                    otros = extract_topics(get_cell(row, otros_col), topic_options["otros_temas_salud_mental"])

                    required_topic_sets = {
                        "SUICIDIO": (suicidio, get_cell(row, suicidio_col)),
                        "VIOLENCIAS": (violencias, get_cell(row, violencias_col)),
                        "ADICCIONES": (adicciones, get_cell(row, adicciones_col)),
                        "OTROS TEMAS DE INTERÉS EN SALUD MENTAL": (otros, get_cell(row, otros_col)),
                    }
                    broken_row = False
                    for section_label, (topics, raw_value) in required_topic_sets.items():
                        if topics != []:
                            continue
                        if clean_text(raw_value) == "":
                            unresolved.append(
                                f"{filename} fila {row_number}: {section_label} está vacío y es obligatorio."
                            )
                        else:
                            unresolved.append(
                                f"{filename} fila {row_number}: {section_label} no se pudo mapear ({clean_text(raw_value)!r})."
                            )
                        broken_row = True
                    if broken_row:
                        continue

                    tema_1 = clean_text(get_cell(row, tema_1_col))
                    tema_2 = clean_text(get_cell(row, tema_2_col))
                    tema_3 = clean_text(get_cell(row, tema_3_col))
                    tema_4 = clean_text(get_cell(row, tema_4_col))
                    justificacion = clean_text(get_cell(row, justificacion_col))
                    created_at = timestamp_to_string(get_cell(row, 0), get_cell(row, 2))

                    payload = {
                        "suicidio": suicidio,
                        "violencias": violencias,
                        "adicciones": adicciones,
                        "otros_temas_salud_mental": otros,
                        "tema_propuesto_1": tema_1,
                        "tema_propuesto_2": tema_2,
                        "tema_propuesto_3": tema_3,
                        "tema_propuesto_4": tema_4,
                        "justificacion_temas": justificacion,
                    }

                    dedupe_key = (
                        str(int(user["id"])),
                        normalize_text(subregion),
                        normalize_text(municipality),
                        created_at,
                        json.dumps(payload, ensure_ascii=False, sort_keys=True),
                    )
                    if dedupe_key in existing_keys:
                        stats["rows_skipped_existing"] += 1
                        continue

                    stats["rows_ready"] += 1

                    if not args.commit:
                        continue

                    cur.execute(
                        insert_sql,
                        (
                            int(user["id"]),
                            str(user["name"]),
                            str(user["email"]),
                            subregion,
                            municipality,
                            int(args.editable),
                            json.dumps(payload, ensure_ascii=False),
                            created_at,
                            created_at,
                        ),
                    )
                    existing_keys.add(dedupe_key)
                    stats["rows_inserted"] += 1

        stats["warnings"] = len(warnings)
        stats["unresolved"] = len(unresolved)

        if unresolved and not args.skip_unresolved:
            connection.rollback()
            print("Hay filas sin resolver. No se aplicaron cambios.", file=sys.stderr)
            for item in unresolved[:20]:
                print("ERROR:", item, file=sys.stderr)
            for item in warnings[:20]:
                print("WARN:", item, file=sys.stderr)
            print(json.dumps(stats, ensure_ascii=False, indent=2))
            return 1

        if args.commit:
            connection.commit()
        else:
            connection.rollback()

        print(json.dumps(stats, ensure_ascii=False, indent=2))
        for item in unresolved[:20]:
            print("ERROR:", item, file=sys.stderr)
        for item in warnings[:20]:
            print("WARN:", item, file=sys.stderr)
        return 0
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
