#!/usr/bin/env python
from __future__ import annotations

import argparse
import datetime as dt
import fnmatch
import hashlib
import json
import os
import re
import sys
import unicodedata
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


ROOT = Path(__file__).resolve().parents[2]
ENV_PATH = ROOT / ".env"
DEFAULT_PATTERN = "Registro_Asistencia*.xlsx"


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    text = text.replace("–", "-").replace("—", "-").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_title_text(value: Any) -> str:
    return clean_text(value).strip()


def get_cell(row: tuple[Any, ...], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def timestamp_to_string(value: Any, fallback: Any = None) -> str:
    if isinstance(value, dt.datetime):
        return value.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min).strftime("%Y-%m-%d %H:%M:%S")
    if fallback is not None:
        return timestamp_to_string(fallback)

    text = clean_text(value)
    if text == "":
        raise ValueError("Fecha vacía.")

    return dt.datetime.fromisoformat(text).replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def date_to_string(value: Any, fallback: Any = None) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if fallback is not None:
        return date_to_string(fallback)

    text = clean_text(value)
    if text == "":
        raise ValueError("Fecha de actividad vacía.")

    return dt.date.fromisoformat(text[:10]).isoformat()


def find_column(headers: list[Any], prefix: str) -> int:
    needle = normalize_text(prefix)
    for idx, header in enumerate(headers):
        if isinstance(header, str) and normalize_text(header).startswith(needle):
            return idx
    return -1


def load_name_aliases(path: str | None) -> dict[str, str]:
    if not path:
        return {}

    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("El archivo de aliases debe ser un JSON objeto {excel_name: target}.")

    aliases: dict[str, str] = {}
    for key, value in raw.items():
        aliases[normalize_text(key)] = str(value)
    return aliases


def build_user_indexes(
    connection: pymysql.connections.Connection,
) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, list[dict[str, Any]]],
    dict[str, list[dict[str, Any]]],
]:
    by_email: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = {}
    by_document: dict[str, list[dict[str, Any]]] = {}

    with connection.cursor() as cur:
        cur.execute("SELECT id, name, email, document_number FROM users")
        for user_id, name, email, document_number in cur.fetchall():
            row = {
                "id": int(user_id),
                "name": str(name),
                "email": str(email),
                "document_number": "" if document_number is None else str(document_number).strip(),
            }
            by_email[normalize_text(email)] = row
            by_name.setdefault(normalize_text(name), []).append(row)
            if row["document_number"] != "":
                by_document.setdefault(normalize_text(row["document_number"]), []).append(row)

    return by_email, by_name, by_document


def resolve_user(
    excel_name: str,
    excel_email: str,
    excel_document: str,
    aliases: dict[str, str],
    by_email: dict[str, dict[str, Any]],
    by_name: dict[str, list[dict[str, Any]]],
    by_document: dict[str, list[dict[str, Any]]],
) -> tuple[dict[str, Any] | None, str]:
    normalized_name = normalize_text(excel_name)
    normalized_email = normalize_text(excel_email)
    normalized_document = normalize_text(excel_document)

    target = aliases.get(normalized_name)
    if target:
        target_normalized = normalize_text(target)
        if target_normalized in by_email:
            return by_email[target_normalized], "alias_email"
        if target_normalized in by_document and len(by_document[target_normalized]) == 1:
            return by_document[target_normalized][0], "alias_document"
        matches = by_name.get(target_normalized, [])
        if len(matches) == 1:
            return matches[0], "alias_name"

    if normalized_email and normalized_email in by_email:
        return by_email[normalized_email], "email"
    if normalized_document and normalized_document in by_document and len(by_document[normalized_document]) == 1:
        return by_document[normalized_document][0], "document"

    matches = by_name.get(normalized_name, [])
    if len(matches) == 1:
        return matches[0], "name"
    if len(matches) > 1:
        return None, "ambiguous"
    return None, "missing"


def collect_files(directory: str, pattern: str) -> list[str]:
    normalized_pattern = normalize_text(pattern)
    files = [
        os.path.join(directory, name)
        for name in os.listdir(directory)
        if fnmatch.fnmatch(normalize_text(name), normalized_pattern)
    ]
    return sorted(files)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa masivamente actividades y asistentes del módulo de asistencia desde Excel."
    )
    parser.add_argument("--file", help="Ruta exacta del xlsx a importar.")
    parser.add_argument("--dir", default=str(Path.home() / "Downloads"), help="Carpeta donde están los xlsx.")
    parser.add_argument("--pattern", default=DEFAULT_PATTERN, help="Patrón de archivos a importar.")
    parser.add_argument("--name-aliases", help="JSON opcional con aliases de nombre o email.")
    parser.add_argument(
        "--skip-unresolved",
        action="store_true",
        help="Importa las filas válidas aunque queden asesores o asistentes sin resolver.",
    )
    parser.add_argument("--commit", action="store_true", help="Inserta en BD. Sin esto solo hace dry-run.")
    return parser.parse_args()


def resolve_input_files(args: argparse.Namespace) -> list[str]:
    if args.file:
        return [args.file]
    return collect_files(args.dir, args.pattern)


def parse_activity_types(raw_value: Any) -> list[str]:
    raw = clean_title_text(raw_value)
    if raw == "":
        return []

    parts = [part.strip() for part in re.split(r"\s+,\s+", raw) if part.strip()]
    if parts == []:
        return [raw]

    seen: OrderedDict[str, str] = OrderedDict()
    for part in parts:
        key = normalize_text(part)
        if key and key not in seen:
            seen[key] = part
    return list(seen.values())


def normalize_group_item(item: str) -> str:
    key = normalize_text(item)
    mappings = {
        "con discapacidad": "Con discapacidad",
        "victima de conflito armado": "Víctima del conflicto armado",
        "victima del conflito armado": "Víctima del conflicto armado",
        "victima del conflicto armado": "Víctima del conflicto armado",
        "victima de conflicto armado": "Víctima del conflicto armado",
        "se considera campesino": "Se considera campesino",
        "¿se considera campesino?": "Se considera campesino",
        "considera que la comunidad donde vive es campesina": "Considera que la comunidad donde vive es campesina",
        "considera que la comunidad en la que vive es campesina": "Considera que la comunidad donde vive es campesina",
        "¿considera que la comunidad en la que vive es campesina?": "Considera que la comunidad donde vive es campesina",
        "ninguno": "Ninguno",
    }
    return mappings.get(key, item.strip())


def parse_group_population(raw_value: Any) -> list[str]:
    raw = clean_title_text(raw_value)
    if raw == "":
        return []

    parts = [part.strip() for part in re.split(r"\s+,\s+", raw) if part.strip()]
    seen: OrderedDict[str, str] = OrderedDict()
    for part in parts:
        normalized = normalize_group_item(part)
        key = normalize_text(normalized)
        if key and key not in seen:
            seen[key] = normalized
    return list(seen.values())


def normalize_zone(raw_value: Any) -> str | None:
    raw = clean_title_text(raw_value)
    if raw == "":
        return None
    key = normalize_text(raw)
    if key == "urbana":
        return "Urbana"
    if key == "rural":
        return "Rural"
    return raw


def normalize_sex(raw_value: Any) -> str | None:
    raw = clean_title_text(raw_value)
    if raw == "":
        return None
    key = normalize_text(raw)
    mappings = {
        "masculino": "Masculino",
        "femenino": "Femenino",
        "no binario": "No binario",
        "transgenero, transexual o travesti": "Transgénero, transexual o travesti",
        "transgenero transexual o travesti": "Transgénero, transexual o travesti",
    }
    return mappings.get(key, raw)


def normalize_etnia(raw_value: Any) -> str | None:
    raw = clean_title_text(raw_value)
    if raw == "":
        return None
    key = normalize_text(raw)
    mappings = {
        "afrodescendiente": "Afrodescendiente",
        "indigena": "Indígena",
        "indigena.": "Indígena",
        "otro": "Otro",
    }
    return mappings.get(key, raw)


def normalize_age(raw_value: Any) -> int | None:
    raw = clean_text(raw_value)
    if raw == "":
        return None
    if re.fullmatch(r"\d+(?:\.0+)?", raw):
        value = int(float(raw))
        if 0 <= value <= 120:
            return value
    return None


def normalized_activity_key(
    subregion: str,
    municipality: str,
    lugar: str,
    advisor_user_id: int,
    activity_date: str,
    activity_types: list[str],
) -> tuple[str, ...]:
    sorted_types = sorted(normalize_text(item) for item in activity_types if normalize_text(item))
    return (
        normalize_text(subregion),
        normalize_text(municipality),
        normalize_text(lugar),
        str(advisor_user_id),
        normalize_text(activity_date),
        *sorted_types,
    )


def build_code_from_key(key: tuple[str, ...]) -> str:
    digest = hashlib.sha1("|".join(key).encode("utf-8")).hexdigest().upper()
    return "ASI-" + digest[:12]


def load_existing_activity_map(
    connection: pymysql.connections.Connection,
) -> tuple[dict[tuple[str, ...], dict[str, Any]], set[str], dict[int, set[str]]]:
    activity_map: dict[tuple[str, ...], dict[str, Any]] = {}
    used_codes: set[str] = set()
    attendee_docs: dict[int, set[str]] = defaultdict(set)

    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT id, code, subregion, municipality, lugar, advisor_user_id, activity_date, actividad_tipos
            FROM asistencia_actividades
            """
        )
        for row in cur.fetchall():
            activity_id = int(row[0])
            code = str(row[1])
            subregion = str(row[2])
            municipality = str(row[3])
            lugar = str(row[4])
            advisor_user_id = int(row[5])
            activity_date = clean_text(row[6])[:10]
            raw_types = row[7]
            decoded = json.loads(raw_types) if raw_types else []
            activity_types = decoded if isinstance(decoded, list) else []
            key = normalized_activity_key(
                subregion,
                municipality,
                lugar,
                advisor_user_id,
                activity_date,
                [str(item) for item in activity_types],
            )
            activity_map[key] = {
                "id": activity_id,
                "code": code,
                "subregion": subregion,
                "municipality": municipality,
                "lugar": lugar,
                "advisor_user_id": advisor_user_id,
                "activity_date": activity_date,
                "actividad_tipos": [str(item) for item in activity_types],
            }
            used_codes.add(code)

        cur.execute("SELECT actividad_id, document_number FROM asistencia_asistentes")
        for actividad_id, document_number in cur.fetchall():
            attendee_docs[int(actividad_id)].add(normalize_text(document_number))

    return activity_map, used_codes, attendee_docs


def ensure_unique_code(base_code: str, used_codes: set[str]) -> str:
    candidate = base_code
    suffix = 1
    while candidate in used_codes:
        extra = f"-{suffix}"
        candidate = (base_code[: 20 - len(extra)] + extra)[:20]
        suffix += 1
    used_codes.add(candidate)
    return candidate


def main() -> int:
    args = parse_args()
    files = resolve_input_files(args)
    if not files:
        print("No se encontraron archivos para importar.", file=sys.stderr)
        return 1

    env = load_env(ENV_PATH)
    aliases = load_name_aliases(args.name_aliases)

    connection = pymysql.connect(
        host=env["DB_HOST"],
        port=int(env.get("DB_PORT", 3306)),
        user=env["DB_USERNAME"],
        password=env.get("DB_PASSWORD", ""),
        database=env["DB_DATABASE"],
        charset="utf8mb4",
        autocommit=False,
    )

    stats = {
        "files": len(files),
        "rows_seen": 0,
        "activities_seen": 0,
        "activities_ready": 0,
        "activities_inserted": 0,
        "activities_skipped_existing": 0,
        "attendees_ready": 0,
        "attendees_inserted": 0,
        "attendees_skipped_existing": 0,
        "attendees_skipped_duplicate_in_file": 0,
        "matched_by_email": 0,
        "matched_by_document": 0,
        "matched_by_name": 0,
        "matched_by_alias_name": 0,
        "matched_by_alias_email": 0,
        "matched_by_alias_document": 0,
        "warnings": 0,
        "unresolved": 0,
    }
    warnings: list[str] = []
    unresolved: list[str] = []

    try:
        by_email, by_name, by_document = build_user_indexes(connection)
        existing_activity_map, used_codes, existing_attendee_docs = load_existing_activity_map(connection)

        insert_activity_sql = """
            INSERT INTO asistencia_actividades (
                code,
                subregion,
                municipality,
                lugar,
                advisor_user_id,
                advisor_name,
                activity_date,
                actividad_tipos,
                status,
                created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        insert_attendee_sql = """
            INSERT INTO asistencia_asistentes (
                actividad_id,
                document_number,
                full_name,
                entity,
                cargo,
                phone,
                email,
                zone,
                sex,
                age,
                etnia,
                etnia_otro,
                grupo_poblacional,
                registered_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        with connection.cursor() as cur:
            for file_path in files:
                filename = os.path.basename(file_path)
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                headers = list(next(rows))

                col_subregion = find_column(headers, "Subregion")
                col_municipality = find_column(headers, "Municipio")
                col_lugar = find_column(headers, "Lugar")
                col_advisor_name = find_column(headers, "Asesor")
                col_advisor_document = find_column(headers, "Cedula_contratista")
                col_advisor_email = find_column(headers, "Correo_contratista")
                col_activity = find_column(headers, "Actividad")
                col_document = find_column(headers, "Documento")
                col_full_name = find_column(headers, "Nombre")
                col_entity = find_column(headers, "Entidad_o_Organizacion")
                col_cargo = find_column(headers, "Cargo")
                col_phone = find_column(headers, "Telefono")
                col_email = find_column(headers, "Correo")
                col_zone = find_column(headers, "Zona")
                col_sex = find_column(headers, "Sexo")
                col_age = find_column(headers, "Edad")
                col_etnia = find_column(headers, "Etnia")
                col_group = find_column(headers, "Grupo_poblacional")
                col_created = find_column(headers, "Fecha")
                col_activity_date = find_column(headers, "Fecha_Actividad")

                required_columns = [
                    col_subregion,
                    col_municipality,
                    col_lugar,
                    col_advisor_name,
                    col_advisor_document,
                    col_advisor_email,
                    col_activity,
                    col_document,
                    col_full_name,
                    col_entity,
                    col_cargo,
                    col_phone,
                    col_email,
                    col_zone,
                    col_sex,
                    col_age,
                    col_etnia,
                    col_group,
                    col_created,
                    col_activity_date,
                ]
                if any(index < 0 for index in required_columns):
                    unresolved.append(f"{filename}: no se pudieron identificar todas las columnas esperadas.")
                    continue

                prepared_activities: dict[tuple[str, ...], dict[str, Any]] = {}

                for row_number, row in enumerate(rows, start=2):
                    stats["rows_seen"] += 1
                    row = tuple(row)

                    advisor_name_excel = clean_title_text(get_cell(row, col_advisor_name))
                    advisor_email_excel = clean_title_text(get_cell(row, col_advisor_email))
                    advisor_document_excel = clean_text(get_cell(row, col_advisor_document))
                    advisor_user, mode = resolve_user(
                        advisor_name_excel,
                        advisor_email_excel,
                        advisor_document_excel,
                        aliases,
                        by_email,
                        by_name,
                        by_document,
                    )
                    if advisor_user is None:
                        unresolved.append(
                            f"{filename} fila {row_number}: no se encontró usuario para asesor "
                            f"{advisor_name_excel!r} ({mode})."
                        )
                        continue
                    stats[f"matched_by_{mode}"] += 1

                    subregion = clean_title_text(get_cell(row, col_subregion)).upper()
                    municipality = clean_title_text(get_cell(row, col_municipality)).upper()
                    lugar = clean_title_text(get_cell(row, col_lugar))
                    activity_types = parse_activity_types(get_cell(row, col_activity))
                    if subregion == "" or municipality == "" or lugar == "" or activity_types == []:
                        unresolved.append(
                            f"{filename} fila {row_number}: faltan datos de actividad (subregión/municipio/lugar/actividad)."
                        )
                        continue

                    if clean_text(get_cell(row, col_activity_date)) == "":
                        warnings.append(
                            f"{filename} fila {row_number}: Fecha_Actividad vacía; se usó Fecha como fallback."
                        )
                    activity_date = date_to_string(get_cell(row, col_activity_date), get_cell(row, col_created))
                    created_at = timestamp_to_string(get_cell(row, col_created), get_cell(row, col_activity_date))
                    activity_key = normalized_activity_key(
                        subregion,
                        municipality,
                        lugar,
                        int(advisor_user["id"]),
                        activity_date,
                        activity_types,
                    )

                    if activity_key not in prepared_activities:
                        prepared_activities[activity_key] = {
                            "key": activity_key,
                            "subregion": subregion,
                            "municipality": municipality,
                            "lugar": lugar,
                            "advisor_user_id": int(advisor_user["id"]),
                            "advisor_name": str(advisor_user["name"]),
                            "activity_date": activity_date,
                            "actividad_tipos": list(activity_types),
                            "created_at": created_at,
                            "attendees": [],
                            "attendee_docs_in_file": set(),
                        }
                    else:
                        prepared_activity = prepared_activities[activity_key]
                        for item in activity_types:
                            if normalize_text(item) not in {
                                normalize_text(existing) for existing in prepared_activity["actividad_tipos"]
                            }:
                                prepared_activity["actividad_tipos"].append(item)
                        if created_at < prepared_activity["created_at"]:
                            prepared_activity["created_at"] = created_at

                    document_number = clean_text(get_cell(row, col_document))
                    full_name = clean_title_text(get_cell(row, col_full_name))
                    if document_number == "":
                        unresolved.append(f"{filename} fila {row_number}: asistente sin documento.")
                        continue
                    if full_name == "":
                        unresolved.append(
                            f"{filename} fila {row_number}: asistente {document_number!r} sin nombre."
                        )
                        continue

                    prepared_activity = prepared_activities[activity_key]
                    document_key = normalize_text(document_number)
                    if document_key in prepared_activity["attendee_docs_in_file"]:
                        stats["attendees_skipped_duplicate_in_file"] += 1
                        warnings.append(
                            f"{filename} fila {row_number}: documento duplicado {document_number!r} dentro de la misma actividad; se omitió."
                        )
                        continue
                    prepared_activity["attendee_docs_in_file"].add(document_key)

                    age = normalize_age(get_cell(row, col_age))
                    if clean_text(get_cell(row, col_age)) != "" and age is None:
                        warnings.append(
                            f"{filename} fila {row_number}: edad inválida {clean_text(get_cell(row, col_age))!r}; se guardó vacía."
                        )

                    attendee = {
                        "document_number": document_number,
                        "full_name": full_name,
                        "entity": clean_title_text(get_cell(row, col_entity)) or None,
                        "cargo": clean_title_text(get_cell(row, col_cargo)) or None,
                        "phone": clean_text(get_cell(row, col_phone)) or None,
                        "email": clean_title_text(get_cell(row, col_email)) or None,
                        "zone": normalize_zone(get_cell(row, col_zone)),
                        "sex": normalize_sex(get_cell(row, col_sex)),
                        "age": age,
                        "etnia": normalize_etnia(get_cell(row, col_etnia)),
                        "etnia_otro": None,
                        "grupo_poblacional": parse_group_population(get_cell(row, col_group)),
                        "registered_at": created_at,
                    }
                    prepared_activity["attendees"].append(attendee)

                stats["activities_seen"] += len(prepared_activities)

                for activity_key, activity in prepared_activities.items():
                    if activity["attendees"] == []:
                        warnings.append(
                            f"{filename}: actividad en {activity['municipality']} / {activity['lugar']} quedó sin asistentes válidos y se omitió."
                        )
                        continue

                    existing_activity = existing_activity_map.get(activity_key)
                    activity_id: int | None = None

                    if existing_activity is not None:
                        activity_id = int(existing_activity["id"])
                        stats["activities_skipped_existing"] += 1
                    else:
                        stats["activities_ready"] += 1
                        if args.commit:
                            code = ensure_unique_code(build_code_from_key(activity_key), used_codes)
                            cur.execute(
                                insert_activity_sql,
                                (
                                    code,
                                    activity["subregion"],
                                    activity["municipality"],
                                    activity["lugar"],
                                    activity["advisor_user_id"],
                                    activity["advisor_name"],
                                    activity["activity_date"],
                                    json.dumps(activity["actividad_tipos"], ensure_ascii=False),
                                    "Pendiente",
                                    activity["created_at"],
                                ),
                            )
                            activity_id = int(cur.lastrowid)
                            existing_activity_map[activity_key] = {
                                "id": activity_id,
                                "code": code,
                            }
                            stats["activities_inserted"] += 1

                    target_existing_docs = (
                        existing_attendee_docs.get(activity_id, set()) if activity_id is not None else set()
                    )

                    for attendee in activity["attendees"]:
                        if normalize_text(attendee["document_number"]) in target_existing_docs:
                            stats["attendees_skipped_existing"] += 1
                            continue

                        stats["attendees_ready"] += 1
                        if not args.commit or activity_id is None:
                            continue

                        cur.execute(
                            insert_attendee_sql,
                            (
                                activity_id,
                                attendee["document_number"],
                                attendee["full_name"],
                                attendee["entity"],
                                attendee["cargo"],
                                attendee["phone"],
                                attendee["email"],
                                attendee["zone"],
                                attendee["sex"],
                                attendee["age"],
                                attendee["etnia"],
                                attendee["etnia_otro"],
                                json.dumps(attendee["grupo_poblacional"], ensure_ascii=False),
                                attendee["registered_at"],
                            ),
                        )
                        existing_attendee_docs.setdefault(activity_id, set()).add(
                            normalize_text(attendee["document_number"])
                        )
                        stats["attendees_inserted"] += 1

        stats["warnings"] = len(warnings)
        stats["unresolved"] = len(unresolved)

        if unresolved and not args.skip_unresolved:
            connection.rollback()
            print("Hay filas sin resolver. No se aplicaron cambios.", file=sys.stderr)
            for item in unresolved[:30]:
                print("ERROR:", item, file=sys.stderr)
            for item in warnings[:30]:
                print("WARN:", item, file=sys.stderr)
            print(json.dumps(stats, ensure_ascii=False, indent=2))
            return 1

        if args.commit:
            connection.commit()
        else:
            connection.rollback()

        print(json.dumps(stats, ensure_ascii=False, indent=2))
        for item in unresolved[:30]:
            print("ERROR:", item, file=sys.stderr)
        for item in warnings[:30]:
            print("WARN:", item, file=sys.stderr)
        return 0
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
