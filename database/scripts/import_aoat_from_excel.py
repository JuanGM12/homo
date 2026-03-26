#!/usr/bin/env python
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


ROOT = Path(__file__).resolve().parents[2]
FORM_PATH = ROOT / "app" / "Views" / "aoat" / "form.php"
ENV_PATH = ROOT / ".env"

MODULE_VALUE_FIELDS = {
    "prev_suicidio",
    "prev_violencias",
    "prev_adicciones",
    "mesa_salud_mental",
    "ppmsmypa",
    "safer",
}


@dataclass
class UserMatch:
    user_id: int
    name: str
    email: str


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("\u2013", " ").replace("\u2014", " ").replace("\u00a0", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_form_options(path: Path) -> dict[str, list[str]]:
    html = path.read_text(encoding="utf-8", errors="replace")
    options: dict[str, list[str]] = {}

    for match in re.finditer(r"<input\b[^>]*>", html, re.IGNORECASE):
        tag = match.group(0)
        name_match = re.search(r'name="([^"]+)"', tag)
        value_match = re.search(r'value="([^"]*)"', tag)
        if not name_match or not value_match:
            continue

        name = name_match.group(1).replace("[]", "")
        value = value_match.group(1)
        if name == "id":
            continue

        options.setdefault(name, [])
        if value not in options[name]:
            options[name].append(value)

    return options


def header_index(headers: list[Any], needle: str) -> int:
    target = normalize_text(needle)
    for idx, header in enumerate(headers):
        if normalize_text(header) == target:
            return idx
    raise KeyError(f"No se encontró la columna '{needle}' en el Excel.")


def first_non_empty(row: tuple[Any, ...], indices: list[int]) -> str:
    for idx in indices:
        value = clean_text(row[idx])
        if value != "":
            return value
    return ""


def as_date(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()

    text = clean_text(value)
    if text == "":
        return ""

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue

    raise ValueError(f"Fecha de actividad no reconocida: {value!r}")


def as_timestamp(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (int, float)):
        if float(value) <= 0:
            raise ValueError("Marca temporal numérica vacía.")
        raise ValueError(f"Marca temporal numérica no soportada: {value!r}")

    text = clean_text(value)
    if text == "":
        raise ValueError("Marca temporal vacía.")

    parsed = dt.datetime.fromisoformat(text)
    return parsed.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def normalize_activity_type(raw: Any) -> str:
    normalized = normalize_text(raw)
    mapping = {
        "asistencia tecnica": "Asistencia técnica",
        "asesoria": "Asesoría",
        "actividad": "Actividad",
    }
    if normalized not in mapping:
        raise ValueError(f"Actividad no reconocida: {raw!r}")
    return mapping[normalized]


def normalize_role(raw: Any) -> tuple[str, str]:
    normalized = normalize_text(raw)
    mapping = {
        "psicologo": ("psicologo", "Psicólogo"),
        "medico": ("medico", "Médico"),
        "abogado": ("abogado", "Abogado"),
        "profesional social": ("profesional social", "Profesional Social"),
    }
    if normalized not in mapping:
        raise ValueError(f"Rol no reconocido: {raw!r}")
    return mapping[normalized]


def normalize_state(raw: Any) -> tuple[str | None, str | None]:
    normalized = normalize_text(raw)
    if normalized == "aprobada":
        return "Aprobada", None
    if normalized == "asignada":
        return "Asignada", None
    if normalized == "devuelta":
        return "Devuelta", None
    if normalized == "sin cargar":
        return "Devuelta", "Sin Cargar en AoAT"
    if normalized == "duplicada":
        return None, None
    raise ValueError(f"Estado no soportado: {raw!r}")


def extract_module_values(raw: Any, allowed_values: list[str]) -> list[str]:
    text = normalize_text(raw)
    if text == "":
        return []

    out: list[str] = []
    module_numbers = {int(n) for n in re.findall(r"\bmodulo\s+(\d+)\b", text)}
    for value in allowed_values:
        normalized_value = normalize_text(value)
        if normalized_value == "no aplica" and "no aplica" in text:
            out.append(value)
            continue

        module_match = re.search(r"\bmodulo\s+(\d+)\b", normalized_value)
        if module_match and int(module_match.group(1)) in module_numbers:
            out.append(value)

    return out


def extract_known_options(raw: Any, allowed_values: list[str]) -> list[str]:
    text = normalize_text(raw)
    if text == "":
        return []

    hits: list[str] = []
    for value in allowed_values:
        if normalize_text(value) in text:
            hits.append(value)

    return hits


def extract_single_option(raw: Any, allowed_values: list[str]) -> str:
    hits = extract_known_options(raw, allowed_values)
    if not hits:
        return ""
    return hits[0]


def build_users_index(connection: pymysql.connections.Connection) -> tuple[dict[str, UserMatch], dict[str, UserMatch]]:
    by_email: dict[str, UserMatch] = {}
    by_name: dict[str, UserMatch] = {}

    with connection.cursor() as cur:
        cur.execute("SELECT id, name, email FROM users")
        for row in cur.fetchall():
            user = UserMatch(user_id=int(row[0]), name=str(row[1]), email=str(row[2]))
            by_email[normalize_text(user.email)] = user
            by_name[normalize_text(user.name)] = user

    return by_email, by_name


def load_existing_keys(connection: pymysql.connections.Connection) -> set[tuple[str, ...]]:
    existing: set[tuple[str, ...]] = set()
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT
                user_id,
                DATE_FORMAT(created_at, '%Y-%m-%d %H:%i:%s') AS created_at_fmt,
                subregion,
                municipality,
                payload
            FROM aoat_records
            """
        )
        for user_id, created_at_fmt, subregion, municipality, payload_json in cur.fetchall():
            payload = json.loads(payload_json) if payload_json else {}
            existing.add(
                make_dedupe_key(
                    int(user_id),
                    str(created_at_fmt),
                    str(subregion or ""),
                    str(municipality or ""),
                    payload,
                )
            )
    return existing


def make_dedupe_key(user_id: int, created_at: str, subregion: str, municipality: str, payload: dict[str, Any]) -> tuple[str, ...]:
    return (
        str(user_id),
        created_at,
        normalize_text(subregion),
        normalize_text(municipality),
        normalize_text(payload.get("aoat_number", "")),
        normalize_text(payload.get("activity_date", "")),
        normalize_text(payload.get("activity_type", "")),
        normalize_text(payload.get("activity_with", "")),
    )


def resolve_user(email_raw: Any, name_raw: Any, by_email: dict[str, UserMatch], by_name: dict[str, UserMatch]) -> tuple[UserMatch | None, str]:
    email_key = normalize_text(email_raw)
    if email_key in by_email:
        return by_email[email_key], "email"

    name_key = normalize_text(name_raw)
    if name_key in by_name:
        return by_name[name_key], "name"

    return None, "missing"


def build_payload(row_map: dict[str, Any], options: dict[str, list[str]]) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "aoat_number": clean_text(row_map.get("aoat_number")),
        "activity_date": as_date(row_map.get("activity_date")),
        "activity_type": normalize_activity_type(row_map.get("activity_type")),
        "activity_with": clean_text(row_map.get("activity_with")),
        "otro_caso": "",
    }

    module_field_headers = {
        "prev_suicidio": [row_map.get("prev_suicidio", "")],
        "prev_violencias": [row_map.get("prev_violencias", ""), row_map.get("prev_violencias_2", "")],
        "prev_adicciones": [row_map.get("prev_adicciones", "")],
        "mesa_salud_mental": [row_map.get("mesa_salud_mental", "")],
        "ppmsmypa": [row_map.get("ppmsmypa", "")],
        "safer": [row_map.get("safer", "")],
    }

    for field, raw_values in module_field_headers.items():
        allowed = options.get(field, [])
        hits: list[str] = []
        for raw in raw_values:
            for hit in extract_module_values(raw, allowed):
                if hit not in hits:
                    hits.append(hit)
        if hits:
            payload[field] = hits

    for field in ("salud_mental", "temas_hospital", "actividad_social"):
        hits = extract_known_options(row_map.get(field, ""), options.get(field, []))
        if hits:
            payload[field] = hits

    proyecto = extract_single_option(row_map.get("proyecto", ""), options.get("proyecto", []))
    if proyecto != "":
        payload["proyecto"] = proyecto

    return payload


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Importa respuestas históricas de AoAT desde un Excel a aoat_records."
    )
    parser.add_argument("--file", required=True, help="Ruta del archivo .xlsx exportado de Forms.")
    parser.add_argument("--commit", action="store_true", help="Inserta en la base de datos. Sin este flag solo hace dry-run.")
    parser.add_argument("--limit", type=int, default=0, help="Procesa solo las primeras N filas de datos.")
    parser.add_argument(
        "--include-duplicadas",
        action="store_true",
        help="Incluye filas con estado 'Duplicada'. Por defecto se omiten.",
    )
    args = parser.parse_args()

    env = load_env(ENV_PATH)
    options = parse_form_options(FORM_PATH)

    workbook = openpyxl.load_workbook(Path(args.file), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))

    header_map = {
        "name": header_index(headers, "Seleccione su nombre"),
        "aoat_number": header_index(headers, "Escriba el número de la AoAT o actividad"),
        "activity_date": header_index(headers, "Fecha de la actividad"),
        "activity_type": header_index(headers, "Actividad que realizó"),
        "state": header_index(headers, "Estado de la AoAT"),
        "role": header_index(headers, "Usted es un:"),
        "prev_suicidio": header_index(headers, "Cualificación temas en prevención del suicidio"),
        "prev_violencias": header_index(headers, "Cualificación temas en prevención de Violencias"),
        "prev_violencias_2": header_index(headers, "Cualificación temas en prevención de Violencias 2"),
        "prev_adicciones": header_index(headers, "Cualificación temas en prevención de Adicciones"),
        "salud_mental": header_index(headers, "Cualificación temas de Salud Mental"),
        "proyecto": header_index(headers, "Proyectos"),
        "temas_hospital": header_index(headers, "Seleccione el tema que dictó en el Hospital del Municipio visitado:"),
        "actividad_social": header_index(headers, "Seleccione la actividad realizada:"),
        "mesa_salud_mental": header_index(headers, "Actualización de la Mesa Municipal de Salud Mental y Prevención de las Adicciones"),
        "ppmsmypa": header_index(headers, "Actualización de la Política Pública Municipal de Salud Y Prevención de las Adiciones (PPMSMYPA)"),
        "safer": header_index(headers, "SAFER"),
        "subregion": header_index(headers, "Seleccione la subregión que visitó"),
        "email": header_index(headers, "Dirección de correo electrónico"),
        "activity_with": header_index(headers, "Con quién realizó la actividad"),
        "observation": header_index(headers, "OBSERVACIÓN"),
        "created_at": header_index(headers, "Marca temporal"),
    }

    municipality_indices = [
        idx
        for idx, header in enumerate(headers)
        if normalize_text(header).startswith("seleccione el municipio visitado")
    ]

    connection = pymysql.connect(
        host=env["DB_HOST"],
        port=int(env.get("DB_PORT", 3306)),
        user=env["DB_USERNAME"],
        password=env.get("DB_PASSWORD", ""),
        database=env["DB_DATABASE"],
        charset="utf8mb4",
        autocommit=False,
    )

    by_email, by_name = build_users_index(connection)
    existing_keys = load_existing_keys(connection)

    insert_sql = """
        INSERT INTO aoat_records (
            user_id,
            professional_name,
            professional_last_name,
            professional_email,
            professional_role,
            profession,
            subregion,
            municipality,
            state,
            audit_observation,
            audit_motive,
            payload,
            created_at,
            updated_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    stats = {
        "rows_seen": 0,
        "rows_ready": 0,
        "rows_inserted": 0,
        "rows_skipped_existing": 0,
        "rows_skipped_duplicada": 0,
        "matched_by_email": 0,
        "matched_by_name": 0,
        "rows_with_warnings": 0,
    }
    warnings: list[str] = []
    unresolved: list[str] = []
    unknown_value_warnings: list[str] = []

    try:
        with connection.cursor() as cur:
            for excel_row_number, row in enumerate(rows, start=2):
                if args.limit and stats["rows_seen"] >= args.limit:
                    break

                stats["rows_seen"] += 1

                raw_state = clean_text(row[header_map["state"]])
                state, audit_motive = normalize_state(raw_state)
                if state is None and normalize_text(raw_state) == "duplicada" and not args.include_duplicadas:
                    stats["rows_skipped_duplicada"] += 1
                    continue
                if state is None:
                    unresolved.append(f"Fila {excel_row_number}: estado no soportado {raw_state!r}.")
                    continue

                user, match_mode = resolve_user(
                    row[header_map["email"]],
                    row[header_map["name"]],
                    by_email,
                    by_name,
                )
                if user is None:
                    unresolved.append(
                        f"Fila {excel_row_number}: no se pudo asociar usuario "
                        f"({clean_text(row[header_map['name']])} / {clean_text(row[header_map['email']])})."
                    )
                    continue
                stats[f"matched_by_{match_mode}"] += 1

                professional_role, profession_label = normalize_role(row[header_map["role"]])
                try:
                    created_at = as_timestamp(row[header_map["created_at"]])
                except ValueError:
                    activity_date_fallback = as_date(row[header_map["activity_date"]])
                    if activity_date_fallback == "":
                        unresolved.append(
                            f"Fila {excel_row_number}: marca temporal inválida y sin fecha de actividad utilizable."
                        )
                        continue
                    created_at = f"{activity_date_fallback} 00:00:00"
                    warnings.append(
                        f"Fila {excel_row_number}: marca temporal inválida; se usó {created_at} como fallback."
                    )
                subregion = clean_text(row[header_map["subregion"]]).upper()
                municipality = first_non_empty(row, municipality_indices).upper()

                if subregion == "" or municipality == "":
                    unresolved.append(f"Fila {excel_row_number}: subregión o municipio vacíos.")
                    continue

                row_map = {
                    "aoat_number": row[header_map["aoat_number"]],
                    "activity_date": row[header_map["activity_date"]],
                    "activity_type": row[header_map["activity_type"]],
                    "activity_with": row[header_map["activity_with"]],
                    "prev_suicidio": row[header_map["prev_suicidio"]],
                    "prev_violencias": row[header_map["prev_violencias"]],
                    "prev_violencias_2": row[header_map["prev_violencias_2"]],
                    "prev_adicciones": row[header_map["prev_adicciones"]],
                    "salud_mental": row[header_map["salud_mental"]],
                    "proyecto": row[header_map["proyecto"]],
                    "temas_hospital": row[header_map["temas_hospital"]],
                    "actividad_social": row[header_map["actividad_social"]],
                    "mesa_salud_mental": row[header_map["mesa_salud_mental"]],
                    "ppmsmypa": row[header_map["ppmsmypa"]],
                    "safer": row[header_map["safer"]],
                }
                payload = build_payload(row_map, options)

                for field in ("salud_mental", "temas_hospital", "actividad_social", "proyecto"):
                    raw = clean_text(row_map.get(field))
                    if raw != "" and field not in payload:
                        unknown_value_warnings.append(
                            f"Fila {excel_row_number}: no se pudo mapear '{field}' desde {raw!r}."
                        )

                for field in MODULE_VALUE_FIELDS:
                    raw_values = row_map.get(field, "")
                    if field == "prev_violencias":
                        raw_values = f"{clean_text(row_map.get('prev_violencias'))} {clean_text(row_map.get('prev_violencias_2'))}".strip()
                    raw_text = clean_text(raw_values)
                    if raw_text != "" and field not in payload:
                        unknown_value_warnings.append(
                            f"Fila {excel_row_number}: no se pudo mapear '{field}' desde {raw_text!r}."
                        )

                audit_observation = clean_text(row[header_map["observation"]]) or None
                if normalize_text(raw_state) == "sin cargar" and audit_observation is None:
                    audit_observation = "Migrado desde Excel con estado original 'Sin Cargar'."
                elif normalize_text(raw_state) not in {"aprobada", "asignada", "devuelta"} and audit_observation is None:
                    audit_observation = f"Migrado desde Excel con estado original '{raw_state}'."

                dedupe_key = make_dedupe_key(user.user_id, created_at, subregion, municipality, payload)
                if dedupe_key in existing_keys:
                    stats["rows_skipped_existing"] += 1
                    continue

                stats["rows_ready"] += 1

                if not args.commit:
                    continue

                data = (
                    user.user_id,
                    user.name,
                    "",
                    user.email,
                    professional_role,
                    profession_label,
                    subregion,
                    municipality,
                    state,
                    audit_observation,
                    audit_motive,
                    json.dumps(payload, ensure_ascii=False),
                    created_at,
                    created_at,
                )
                cur.execute(insert_sql, data)
                existing_keys.add(dedupe_key)
                stats["rows_inserted"] += 1

        if unresolved or unknown_value_warnings:
            connection.rollback()
            stats["rows_with_warnings"] = len(warnings) + len(unresolved) + len(unknown_value_warnings)
            print("Se encontraron incidencias. No se aplicaron cambios.", file=sys.stderr)
            for message in unresolved[:20]:
                print("ERROR:", message, file=sys.stderr)
            for message in unknown_value_warnings[:20]:
                print("WARN:", message, file=sys.stderr)
            for message in warnings[:20]:
                print("WARN:", message, file=sys.stderr)
            print(json.dumps(stats, ensure_ascii=False, indent=2))
            return 1

        if args.commit:
            connection.commit()
        else:
            connection.rollback()

        stats["rows_with_warnings"] = len(warnings)
        print(json.dumps(stats, ensure_ascii=False, indent=2))
        for message in warnings[:20]:
            print("WARN:", message, file=sys.stderr)
        if stats["rows_skipped_duplicada"] > 0:
            print(
                "Nota: las filas con estado 'Duplicada' se omitieron por defecto. Usa --include-duplicadas si quieres evaluarlas.",
                file=sys.stderr,
            )
        return 0
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
