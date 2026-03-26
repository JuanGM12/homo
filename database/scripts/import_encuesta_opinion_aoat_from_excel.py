#!/usr/bin/env python
from __future__ import annotations

import argparse
import datetime as dt
import fnmatch
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


ROOT = Path(__file__).resolve().parents[2]
ENV_PATH = ROOT / ".env"
DEFAULT_PATTERN = "Encuesta de Opinion del Usuario de la AoAT*.xlsx"


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    text = text.replace("–", "-").replace("—", "-").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def get_cell(row: tuple[Any, ...], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def first_non_empty(values: list[Any]) -> str:
    for value in values:
        text = clean_text(value)
        if text != "":
            return text
    return ""


def timestamp_to_string(value: Any, fallback: Any = None) -> str:
    if isinstance(value, dt.datetime):
        return value.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min).strftime("%Y-%m-%d %H:%M:%S")
    if fallback is not None:
        return timestamp_to_string(fallback)

    text = clean_text(value)
    if text == "":
        raise ValueError("Marca temporal vacía.")

    return dt.datetime.fromisoformat(text).replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def date_to_string(value: Any, fallback: Any = None) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if fallback is not None:
        return date_to_string(fallback)

    text = clean_text(value)
    if text == "":
        raise ValueError("Fecha vacía.")

    iso_candidate = text[:10]
    try:
        return dt.date.fromisoformat(iso_candidate).isoformat()
    except ValueError:
        pass

    slash_match = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{1,4})", text)
    if slash_match:
        month = int(slash_match.group(1))
        day = int(slash_match.group(2))
        year = int(slash_match.group(3))
        if year < 1000:
            year = 2000 + (year % 100)
        return dt.date(year, month, day).isoformat()

    raise ValueError(f"Fecha inválida: {text!r}")


def find_column(headers: list[Any], prefix: str) -> int:
    needle = normalize_text(prefix)
    for idx, header in enumerate(headers):
        if isinstance(header, str) and normalize_text(header).startswith(needle):
            return idx
    return -1


def find_columns(headers: list[Any], prefix: str) -> list[int]:
    needle = normalize_text(prefix)
    indices: list[int] = []
    for idx, header in enumerate(headers):
        if isinstance(header, str) and normalize_text(header).startswith(needle):
            indices.append(idx)
    return indices


def load_name_aliases(path: str | None) -> dict[str, str]:
    if not path:
        return {}

    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("El archivo de aliases debe ser un JSON objeto {excel_name: target}.")

    aliases: dict[str, str] = {}
    for key, value in raw.items():
        aliases[normalize_text(key)] = str(value)
    return aliases


def build_user_indexes(
    connection: pymysql.connections.Connection,
) -> dict[str, list[dict[str, Any]]]:
    by_name: dict[str, list[dict[str, Any]]] = {}
    with connection.cursor() as cur:
        cur.execute("SELECT id, name, email FROM users")
        for user_id, name, email in cur.fetchall():
            row = {"id": int(user_id), "name": str(name), "email": str(email)}
            by_name.setdefault(normalize_text(name), []).append(row)
    return by_name


def resolve_user(
    excel_name: str,
    aliases: dict[str, str],
    by_name: dict[str, list[dict[str, Any]]],
) -> tuple[dict[str, Any] | None, str]:
    normalized_name = normalize_text(excel_name)
    target = aliases.get(normalized_name)
    if target:
        matches = by_name.get(normalize_text(target), [])
        if len(matches) == 1:
            return matches[0], "alias_name"

    matches = by_name.get(normalized_name, [])
    if len(matches) == 1:
        return matches[0], "name"
    if len(matches) > 1:
        return None, "ambiguous"
    return None, "missing"


def normalize_score(value: Any) -> int | None:
    raw = clean_text(value)
    if raw == "":
        return None
    if re.fullmatch(r"[1-5](?:\.0+)?", raw):
        return int(float(raw))
    return None


def load_existing_keys(connection: pymysql.connections.Connection) -> set[tuple[str, ...]]:
    keys: set[tuple[str, ...]] = set()
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT advisor_user_id, actividad, lugar, activity_date, subregion, municipality,
                   score_objetivos, score_claridad, score_pertinencia, score_ayudas, score_relacion, score_puntualidad,
                   comments, created_at
            FROM encuesta_opinion_aoat
            """
        )
        for row in cur.fetchall():
            keys.add(tuple(normalize_text(item) for item in row))
    return keys


def collect_files(directory: str, pattern: str) -> list[str]:
    normalized_pattern = normalize_text(pattern)
    files = [
        os.path.join(directory, name)
        for name in os.listdir(directory)
        if fnmatch.fnmatch(normalize_text(name), normalized_pattern)
    ]
    return sorted(files)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa masivamente la encuesta de opinión AoAT desde Excel."
    )
    parser.add_argument("--file", help="Ruta exacta del xlsx a importar.")
    parser.add_argument("--dir", default=str(Path.home() / "Downloads"), help="Carpeta donde están los xlsx.")
    parser.add_argument("--pattern", default=DEFAULT_PATTERN, help="Patrón de archivos a importar.")
    parser.add_argument("--name-aliases", help="JSON opcional con aliases de nombre.")
    parser.add_argument(
        "--skip-unresolved",
        action="store_true",
        help="Importa las filas válidas aunque queden asesores sin resolver.",
    )
    parser.add_argument("--commit", action="store_true", help="Inserta en BD. Sin esto solo hace dry-run.")
    return parser.parse_args()


def resolve_input_files(args: argparse.Namespace) -> list[str]:
    if args.file:
        return [args.file]
    return collect_files(args.dir, args.pattern)


def main() -> int:
    args = parse_args()
    files = resolve_input_files(args)
    if not files:
        print("No se encontraron archivos para importar.", file=sys.stderr)
        return 1

    env = load_env(ENV_PATH)
    aliases = load_name_aliases(args.name_aliases)

    connection = pymysql.connect(
        host=env["DB_HOST"],
        port=int(env.get("DB_PORT", 3306)),
        user=env["DB_USERNAME"],
        password=env.get("DB_PASSWORD", ""),
        database=env["DB_DATABASE"],
        charset="utf8mb4",
        autocommit=False,
    )

    stats = {
        "files": len(files),
        "rows_seen": 0,
        "rows_ready": 0,
        "rows_inserted": 0,
        "rows_skipped_existing": 0,
        "matched_by_name": 0,
        "matched_by_alias_name": 0,
        "warnings": 0,
        "unresolved": 0,
    }
    warnings: list[str] = []
    unresolved: list[str] = []

    try:
        by_name = build_user_indexes(connection)
        existing_keys = load_existing_keys(connection)

        insert_sql = """
            INSERT INTO encuesta_opinion_aoat (
                advisor_user_id,
                advisor_name,
                actividad,
                lugar,
                activity_date,
                subregion,
                municipality,
                score_objetivos,
                score_claridad,
                score_pertinencia,
                score_ayudas,
                score_relacion,
                score_puntualidad,
                comments,
                created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        with connection.cursor() as cur:
            for file_path in files:
                filename = os.path.basename(file_path)
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                headers = list(next(rows))

                col_name = find_column(headers, "Seleccione el nombre del asesor")
                col_activity = find_column(headers, "Actividad")
                col_place = find_column(headers, "Lugar")
                col_date = find_column(headers, "Fecha")
                col_subregion = find_column(headers, "Seleccione la subregión")
                municipality_cols = find_columns(headers, "Seleccione el municipio de pertenencia")
                score_cols = list(range(15, 21))
                col_comments = find_column(headers, "Recomendaciones o comentarios")

                required_columns = [col_name, col_activity, col_place, col_date, col_subregion, col_comments]
                if municipality_cols == [] or any(index < 0 for index in required_columns):
                    unresolved.append(f"{filename}: no se pudieron identificar todas las columnas esperadas.")
                    continue

                for row_number, row in enumerate(rows, start=2):
                    stats["rows_seen"] += 1
                    row = tuple(row)

                    excel_name = clean_text(get_cell(row, col_name))
                    user, mode = resolve_user(excel_name, aliases, by_name)
                    if user is None:
                        unresolved.append(
                            f"{filename} fila {row_number}: no se encontró usuario para {excel_name!r} ({mode})."
                        )
                        continue
                    stats[f"matched_by_{mode}"] += 1

                    actividad = clean_text(get_cell(row, col_activity))
                    lugar = clean_text(get_cell(row, col_place))
                    activity_date = date_to_string(get_cell(row, col_date))
                    created_at = timestamp_to_string(get_cell(row, 0), get_cell(row, col_date))
                    subregion = clean_text(get_cell(row, col_subregion)).upper()
                    municipality = first_non_empty([get_cell(row, idx) for idx in municipality_cols]).upper()
                    comments = clean_text(get_cell(row, col_comments)) or None

                    if actividad == "" or lugar == "" or subregion == "" or municipality == "":
                        unresolved.append(
                            f"{filename} fila {row_number}: faltan datos obligatorios de actividad/lugar/subregión/municipio."
                        )
                        continue

                    scores: list[int] = []
                    bad_score = False
                    for score_idx in score_cols:
                        score = normalize_score(get_cell(row, score_idx))
                        if score is None:
                            unresolved.append(
                                f"{filename} fila {row_number}: puntaje inválido en columna {score_idx + 1}."
                            )
                            bad_score = True
                            break
                        scores.append(score)
                    if bad_score:
                        continue

                    dedupe_key = tuple(
                        normalize_text(item)
                        for item in (
                            int(user["id"]),
                            str(user["name"]),
                            actividad,
                            lugar,
                            activity_date,
                            subregion,
                            municipality,
                            scores[0],
                            scores[1],
                            scores[2],
                            scores[3],
                            scores[4],
                            scores[5],
                            comments or "",
                            created_at,
                        )
                    )
                    if dedupe_key in existing_keys:
                        stats["rows_skipped_existing"] += 1
                        continue

                    stats["rows_ready"] += 1

                    if not args.commit:
                        continue

                    cur.execute(
                        insert_sql,
                        (
                            int(user["id"]),
                            str(user["name"]),
                            actividad,
                            lugar,
                            activity_date,
                            subregion,
                            municipality,
                            scores[0],
                            scores[1],
                            scores[2],
                            scores[3],
                            scores[4],
                            scores[5],
                            comments,
                            created_at,
                        ),
                    )
                    existing_keys.add(dedupe_key)
                    stats["rows_inserted"] += 1

        stats["warnings"] = len(warnings)
        stats["unresolved"] = len(unresolved)

        if unresolved and not args.skip_unresolved:
            connection.rollback()
            print("Hay filas sin resolver. No se aplicaron cambios.", file=sys.stderr)
            for item in unresolved[:20]:
                print("ERROR:", item, file=sys.stderr)
            for item in warnings[:20]:
                print("WARN:", item, file=sys.stderr)
            print(json.dumps(stats, ensure_ascii=False, indent=2))
            return 1

        if args.commit:
            connection.commit()
        else:
            connection.rollback()

        print(json.dumps(stats, ensure_ascii=False, indent=2))
        for item in unresolved[:20]:
            print("ERROR:", item, file=sys.stderr)
        for item in warnings[:20]:
            print("WARN:", item, file=sys.stderr)
        return 0
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
