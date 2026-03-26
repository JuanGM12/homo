#!/usr/bin/env python
from __future__ import annotations

import argparse
import datetime as dt
import fnmatch
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


ROOT = Path(__file__).resolve().parents[2]
ENV_PATH = ROOT / ".env"
FORM_PATH = ROOT / "app" / "Views" / "planeacion" / "form.php"
MONTH_KEYS = [
    ("enero", "Enero"),
    ("febrero", "Febrero"),
    ("marzo", "Marzo"),
    ("abril", "Abril"),
    ("mayo", "Mayo"),
    ("junio", "Junio"),
    ("julio", "Julio"),
    ("agosto", "Agosto"),
    ("septiembre", "Septiembre"),
    ("octubre", "Octubre"),
    ("noviembre", "Noviembre"),
    ("diciembre", "Diciembre"),
]
DEFAULT_PATTERN = "Planeacion anual capacitaciones municipales*.xlsx"


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    text = text.replace("–", "-").replace("—", "-").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def parse_php_string_list(var_name: str, text: str) -> list[str]:
    match = re.search(rf"\${var_name}\s*=\s*\[(.*?)\];", text, re.S)
    if not match:
        return []

    block = match.group(1)
    values: list[str] = []
    for raw in re.findall(r"'((?:\\'|[^'])*)'", block):
        values.append(raw.replace("\\'", "'"))

    return values


def load_topic_options(form_path: Path) -> dict[str, list[str]]:
    text = form_path.read_text(encoding="utf-8", errors="replace")
    return {
        "abogado": parse_php_string_list("topicOptionsAbogado", text),
        "medico": parse_php_string_list("topicOptionsMedico", text),
        "psicologo": parse_php_string_list("topicOptionsPsicologo", text),
    }


def detect_role_from_filename(filename: str) -> str:
    normalized = normalize_text(filename)
    if "abogados" in normalized:
        return "abogado"
    if "medicos" in normalized:
        return "medico"
    if "psicologos" in normalized:
        return "psicologo"
    raise ValueError(f"No se pudo detectar el rol desde el archivo: {filename}")


def timestamp_to_string(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min).strftime("%Y-%m-%d %H:%M:%S")

    text = clean_text(value)
    if text == "":
        raise ValueError("Marca temporal vacía.")

    return dt.datetime.fromisoformat(text).replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def header_indices(headers: list[Any], prefix: str) -> list[int]:
    out: list[int] = []
    for idx, header in enumerate(headers):
        if isinstance(header, str) and header.startswith(prefix):
            out.append(idx)
    return out


def first_non_empty(values: list[Any]) -> str:
    for value in values:
        text = clean_text(value)
        if text != "":
            return text
    return ""


def load_name_aliases(path: str | None) -> dict[str, str]:
    if not path:
        return {}

    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError("El archivo de aliases debe ser un JSON objeto {excel_name: target}.")

    aliases: dict[str, str] = {}
    for key, value in raw.items():
        aliases[normalize_text(key)] = str(value)
    return aliases


def build_user_indexes(connection: pymysql.connections.Connection) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    by_email: dict[str, dict[str, Any]] = {}
    by_name: dict[str, list[dict[str, Any]]] = {}

    with connection.cursor() as cur:
        cur.execute("SELECT id, name, email FROM users")
        for user_id, name, email in cur.fetchall():
            row = {"id": int(user_id), "name": str(name), "email": str(email)}
            by_email[normalize_text(email)] = row
            by_name.setdefault(normalize_text(name), []).append(row)

    return by_email, by_name


def resolve_user(
    excel_name: str,
    aliases: dict[str, str],
    by_email: dict[str, dict[str, Any]],
    by_name: dict[str, list[dict[str, Any]]],
) -> tuple[dict[str, Any] | None, str]:
    normalized_name = normalize_text(excel_name)
    target = aliases.get(normalized_name)
    if target:
        target_normalized = normalize_text(target)
        if target_normalized in by_email:
            return by_email[target_normalized], "alias_email"
        matches = by_name.get(target_normalized, [])
        if len(matches) == 1:
            return matches[0], "alias_name"

    matches = by_name.get(normalized_name, [])
    if len(matches) == 1:
        return matches[0], "name"
    if len(matches) > 1:
        return None, "ambiguous"
    return None, "missing"


def extract_topics(cell_value: Any, allowed_topics: list[str]) -> list[str]:
    text = normalize_text(cell_value)
    if text == "":
        return []

    hits: list[str] = []
    for topic in allowed_topics:
        if normalize_text(topic) in text:
            hits.append(topic)

    return hits


def should_warn_overflow(topics_text: str, population_text: str) -> bool:
    normalized_topics = normalize_text(topics_text)
    normalized_population = normalize_text(population_text)
    if normalized_topics in ("", "no aplica"):
        return False
    if normalized_population in ("", "no aplica"):
        return False
    return True


def build_payload(
    row: tuple[Any, ...],
    topic_columns: list[int],
    population_columns: list[int],
    allowed_topics: list[str],
    warnings: list[str],
    file_label: str,
    row_number: int,
) -> dict[str, dict[str, Any]]:
    payload: dict[str, dict[str, Any]] = {}
    usable_pairs = min(len(MONTH_KEYS), len(topic_columns), len(population_columns))

    for month_index in range(usable_pairs):
        month_key, month_label = MONTH_KEYS[month_index]
        raw_topics = row[topic_columns[month_index]]
        raw_population = row[population_columns[month_index]]
        topics = extract_topics(raw_topics, allowed_topics)
        population = clean_text(raw_population)
        normalized_cell = normalize_text(raw_topics)

        if topics == [] and normalized_cell not in ("",):
            warnings.append(
                f"{file_label} fila {row_number}: mes {month_label} no se pudo mapear ({clean_text(raw_topics)!r})."
            )

        if topics or population != "":
            payload[month_key] = {
                "label": month_label,
                "topics": topics,
                "population": population,
            }

    if len(topic_columns) > len(MONTH_KEYS):
        for overflow_index in range(len(MONTH_KEYS), len(topic_columns)):
            overflow_topics = clean_text(row[topic_columns[overflow_index]])
            overflow_population = clean_text(
                row[population_columns[overflow_index]] if overflow_index < len(population_columns) else ""
            )
            if should_warn_overflow(overflow_topics, overflow_population):
                warnings.append(
                    f"{file_label} fila {row_number}: se ignoró una columna extra fuera de los 12 meses "
                    f"({overflow_topics!r} / {overflow_population!r})."
                )

    return payload


def load_existing_keys(connection: pymysql.connections.Connection) -> set[tuple[str, ...]]:
    keys: set[tuple[str, ...]] = set()
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT user_id, plan_year, subregion, municipality, created_at, payload
            FROM training_plans
            """
        )
        for user_id, plan_year, subregion, municipality, created_at, payload_json in cur.fetchall():
            payload = json.loads(payload_json) if payload_json else {}
            keys.add(
                (
                    str(int(user_id)),
                    str(int(plan_year)),
                    normalize_text(subregion),
                    normalize_text(municipality),
                    clean_text(created_at),
                    json.dumps(payload, ensure_ascii=False, sort_keys=True),
                )
            )
    return keys


def collect_files(directory: str, pattern: str) -> list[str]:
    normalized_pattern = normalize_text(pattern)
    files = [
        os.path.join(directory, name)
        for name in os.listdir(directory)
        if fnmatch.fnmatch(normalize_text(name), normalized_pattern)
    ]
    return sorted(files)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Importa masivamente planeaciones anuales desde Excel a training_plans."
    )
    parser.add_argument("--dir", default=str(Path.home() / "Downloads"), help="Carpeta donde están los xlsx.")
    parser.add_argument("--pattern", default=DEFAULT_PATTERN, help="Patrón de archivos a importar.")
    parser.add_argument("--plan-year", type=int, default=dt.date.today().year, help="Año de la planeación.")
    parser.add_argument("--editable", type=int, default=1, choices=[0, 1], help="Valor del campo editable.")
    parser.add_argument("--name-aliases", help="JSON opcional con aliases de nombre o email.")
    parser.add_argument(
        "--skip-unresolved",
        action="store_true",
        help="Importa las filas válidas aunque queden profesionales sin resolver.",
    )
    parser.add_argument("--commit", action="store_true", help="Inserta en BD. Sin esto solo hace dry-run.")
    args = parser.parse_args()

    files = collect_files(args.dir, args.pattern)
    if not files:
        print("No se encontraron archivos para importar.", file=sys.stderr)
        return 1

    env = load_env(ENV_PATH)
    topic_options = load_topic_options(FORM_PATH)
    aliases = load_name_aliases(args.name_aliases)

    connection = pymysql.connect(
        host=env["DB_HOST"],
        port=int(env.get("DB_PORT", 3306)),
        user=env["DB_USERNAME"],
        password=env.get("DB_PASSWORD", ""),
        database=env["DB_DATABASE"],
        charset="utf8mb4",
        autocommit=False,
    )

    stats = {
        "files": len(files),
        "rows_seen": 0,
        "rows_ready": 0,
        "rows_inserted": 0,
        "rows_skipped_existing": 0,
        "matched_by_name": 0,
        "matched_by_alias_name": 0,
        "matched_by_alias_email": 0,
        "warnings": 0,
        "unresolved": 0,
    }
    warnings: list[str] = []
    unresolved: list[str] = []

    try:
        by_email, by_name = build_user_indexes(connection)
        existing_keys = load_existing_keys(connection)

        insert_sql = """
            INSERT INTO training_plans (
                user_id,
                professional_name,
                professional_email,
                professional_role,
                subregion,
                municipality,
                plan_year,
                editable,
                payload,
                created_at,
                updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        with connection.cursor() as cur:
            for file_path in files:
                filename = os.path.basename(file_path)
                role = detect_role_from_filename(filename)
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                headers = list(next(rows))

                municipality_columns = header_indices(headers, "Seleccione el municipio visitado")
                topic_columns = header_indices(headers, "Seleccione los Temas / Módulos de interés a desarrollar durante el mes")
                population_columns = header_indices(headers, "Población objetivo")
                if not municipality_columns or not topic_columns or not population_columns:
                    unresolved.append(f"{filename}: no se pudieron identificar todas las columnas esperadas.")
                    continue

                for row_number, row in enumerate(rows, start=2):
                    stats["rows_seen"] += 1
                    excel_name = clean_text(row[1])
                    user, mode = resolve_user(excel_name, aliases, by_email, by_name)
                    if user is None:
                        unresolved.append(f"{filename} fila {row_number}: no se encontró usuario para {excel_name!r} ({mode}).")
                        continue
                    stats[f"matched_by_{mode}"] += 1

                    subregion = first_non_empty([row[2]]).upper()
                    municipality = first_non_empty([row[idx] for idx in municipality_columns]).upper()
                    if subregion == "" or municipality == "":
                        unresolved.append(f"{filename} fila {row_number}: subregión o municipio vacíos.")
                        continue

                    created_at = timestamp_to_string(row[0])
                    payload = build_payload(
                        row,
                        topic_columns,
                        population_columns,
                        topic_options[role],
                        warnings,
                        filename,
                        row_number,
                    )

                    if payload == {}:
                        warnings.append(f"{filename} fila {row_number}: sin meses diligenciados, se omitió.")
                        continue

                    dedupe_key = (
                        str(int(user["id"])),
                        str(args.plan_year),
                        normalize_text(subregion),
                        normalize_text(municipality),
                        created_at,
                        json.dumps(payload, ensure_ascii=False, sort_keys=True),
                    )
                    if dedupe_key in existing_keys:
                        stats["rows_skipped_existing"] += 1
                        continue

                    stats["rows_ready"] += 1

                    if not args.commit:
                        continue

                    cur.execute(
                        insert_sql,
                        (
                            int(user["id"]),
                            str(user["name"]),
                            str(user["email"]),
                            role,
                            subregion,
                            municipality,
                            int(args.plan_year),
                            int(args.editable),
                            json.dumps(payload, ensure_ascii=False),
                            created_at,
                            created_at,
                        ),
                    )
                    existing_keys.add(dedupe_key)
                    stats["rows_inserted"] += 1

        stats["warnings"] = len(warnings)
        stats["unresolved"] = len(unresolved)

        if unresolved and not args.skip_unresolved:
            connection.rollback()
            print("Hay filas sin resolver. No se aplicaron cambios.", file=sys.stderr)
            for item in unresolved[:20]:
                print("ERROR:", item, file=sys.stderr)
            for item in warnings[:20]:
                print("WARN:", item, file=sys.stderr)
            print(json.dumps(stats, ensure_ascii=False, indent=2))
            return 1

        if args.commit:
            connection.commit()
        else:
            connection.rollback()

        print(json.dumps(stats, ensure_ascii=False, indent=2))
        for item in unresolved[:20]:
            print("ERROR:", item, file=sys.stderr)
        for item in warnings[:20]:
            print("WARN:", item, file=sys.stderr)
        return 0
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
